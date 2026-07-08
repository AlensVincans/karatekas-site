from __future__ import annotations

import csv
import html
import json
import re
import shutil
import sys
import urllib.parse
from collections import Counter, defaultdict
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
EXCEL_SOURCE = Path(r"C:/Users/alenv/Downloads/Sajt_karatekas kihon_price_list.xlsx")
WOO_SOURCE = Path(r"C:/Users/alenv/Downloads/wc-product-export-23-6-2026-1782247894297.csv")
UPLOADS_ROOT = Path(r"C:/Users/alenv/Downloads/uploads")
PRODUCT_IMAGE_DIR = ROOT / "public" / "product-images"
TS_PATH = ROOT / "lib" / "karate-products.ts"
KNOWN_BRANDS = {
    "Arawaza",
    "Best Sport",
    "Century",
    "Kihon",
    "Punok",
    "Sentei",
    "Smai",
    "Tokaido",
}
CATEGORY_ORDER = {
    "Kimono": 10,
    "Protective Equipment": 20,
    "Belts": 30,
    "Punching bags": 40,
    "Tatami": 50,
    "Accessories": 60,
    "WUKF": 70,
    "Equipment": 80,
}

sys.stdout.reconfigure(encoding="utf-8")


def clean(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    text = str(value).replace("\r\n", "\n").replace("\r", "\n").strip()
    return re.sub(r"[ \t]+", " ", html.unescape(text))


def strip_html(value: str) -> str:
    text = re.sub(r"<[^>]+>", " ", clean(value))
    return re.sub(r"\s+", " ", text).strip()


def number(value: Any, default: float = 0) -> float:
    if value is None or value == "":
        return default
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).replace(",", ".").strip()
    try:
        return float(text)
    except ValueError:
        return default


def slug(value: str) -> str:
    value = clean(value).lower()
    value = re.sub(r"[^a-z0-9а-яё]+", "-", value, flags=re.IGNORECASE)
    value = value.strip("-")
    return value or "item"


def safe_asset_name(value: str) -> str:
    name = urllib.parse.unquote(value)
    name = re.sub(r"[^a-zA-Z0-9._-]+", "-", name).strip("-")
    return name.lower() or "image.jpg"


def ts(value: Any) -> str:
    return json.dumps(value, ensure_ascii=False)


def category_name(raw: str) -> str:
    key = clean(raw).lower()
    if "kimono" in key or "gi" in key:
        return "Kimono"
    if "protect" in key or "guard" in key:
        return "Protective Equipment"
    if "belt" in key:
        return "Belts"
    if "punching" in key or "bag" in key or "dummy" in key:
        return "Punching bags"
    if "tatami" in key:
        return "Tatami"
    if "wukf" in key:
        return "WUKF"
    if "accessor" in key:
        return "Accessories"
    if "music box" in key:
        return "Accessories"
    return clean(raw) or "Equipment"


def colour_name(raw: str) -> str:
    mapping = {
        "white": "White",
        "black": "Black",
        "blue": "Blue",
        "red": "Red",
        "pink": "Pink",
        "green": "Green",
        "yellow": "Yellow",
        "orange": "Orange",
        "grey": "Grey",
        "gray": "Grey",
        "skin colour": "Skin",
        "skin color": "Skin",
        "no colour": "",
        "no color": "",
    }
    key = clean(raw).lower()
    return mapping.get(key, clean(raw))


def split_values(value: str) -> list[str]:
    parts = [clean(part) for part in re.split(r"\s*,\s*|\s*\|\s*", clean(value)) if clean(part)]
    return parts or [""]


def find_excel_header(rows: list[tuple[Any, ...]]) -> tuple[int, list[str]] | None:
    for index, row in enumerate(rows[:20]):
        headers = [clean(cell) for cell in row]
        lowered = {header.lower() for header in headers}
        if "brand" in lowered and "title" in lowered:
            return index, headers
    return None


def row_value(row: tuple[Any, ...], headers: list[str], name: str) -> Any:
    wanted = name.lower()
    for index, header in enumerate(headers):
        if header.lower() == wanted:
            return row[index] if index < len(row) else None
    return None


def unique_value(base: str, seen: Counter[str]) -> str:
    base = clean(base) or "item"
    seen[base] += 1
    if seen[base] == 1:
        return base
    return f"{base}-{seen[base]}"


def parse_attrs(row: dict[str, Any]) -> dict[str, list[str]]:
    attrs: dict[str, list[str]] = {}
    for index in (1, 2):
        name = clean(row.get(f"Atribūta {index} nosaukums"))
        value = clean(row.get(f"Atribūta {index} vērtība (-s)"))
        if not name or not value:
            continue
        key = name.lower()
        if "color" in key or "colour" in key:
            attrs["color"] = [colour_name(part) for part in split_values(value)]
        elif "size" in key:
            attrs["size"] = split_values(value)
        else:
            attrs[name.lower()] = split_values(value)
    return attrs


def import_excel_rows() -> list[dict[str, Any]]:
    if not EXCEL_SOURCE.exists():
        return []

    wb = load_workbook(EXCEL_SOURCE, data_only=True)
    imported_at = datetime.now(timezone.utc).isoformat()
    items: list[dict[str, Any]] = []
    seen_skus: Counter[str] = Counter()

    for ws in wb.worksheets:
        rows = list(ws.iter_rows(values_only=True))
        header_match = find_excel_header(rows)
        if not header_match:
            continue

        header_index, headers = header_match
        for excel_row_number, row in enumerate(rows[header_index + 1 :], start=header_index + 2):
            if not any(cell is not None and clean(cell) != "" for cell in row):
                continue

            brand = clean(row_value(row, headers, "Brand")) or ws.title
            group = clean(row_value(row, headers, "Group")) or "Equipment"
            title = clean(row_value(row, headers, "Title"))
            description = clean(row_value(row, headers, "Description"))
            size = clean(row_value(row, headers, "Size"))
            colour = colour_name(clean(row_value(row, headers, "Colour")))
            club_price = number(row_value(row, headers, "Club Price"))
            retail_price = number(row_value(row, headers, "I-K Price"))
            stock = int(round(number(row_value(row, headers, "Stock"))))

            if not title:
                continue

            if retail_price <= 0 and club_price > 0:
                retail_price = round(club_price * 1.28, 2)
            if club_price <= 0 and retail_price > 0:
                club_price = round(retail_price * 0.78, 2)

            raw_sku = clean(row_value(row, headers, "Product code/ SKU"))
            generated = f"{slug(brand)[:4].upper()}-{slug(title)[:18].upper()}-{slug(size or colour or str(excel_row_number)).upper()}"
            sku = unique_value(raw_sku or generated, seen_skus)

            images = [clean(row_value(row, headers, f"img{i}")) for i in range(1, 5)]
            items.append(
                {
                    "source_sheet": f"excel:{ws.title}",
                    "brand": brand,
                    "product_group": group,
                    "category": category_name(group),
                    "sku": sku,
                    "title": title,
                    "description": description,
                    "stock": max(0, stock),
                    "size": size,
                    "colour": colour,
                    "club_price": club_price,
                    "retail_price": retail_price,
                    "images": images,
                    "image_urls": [],
                    "imported_at": imported_at,
                }
            )

    return items


def image_urls(row: dict[str, Any]) -> list[str]:
    raw = clean(row.get("Attēli"))
    if not raw:
        return []
    return [item.strip() for item in raw.split(",") if item.strip() and item.strip() != "0"]


def upload_index() -> dict[str, Path]:
    files = [path for path in UPLOADS_ROOT.rglob("*") if path.is_file()]
    return {path.name.lower(): path for path in files}


def url_candidates(url: str) -> list[str]:
    name = urllib.parse.unquote(Path(urllib.parse.urlparse(url).path).name)
    stem = Path(name).stem
    suffix = Path(name).suffix
    return [
        name,
        re.sub(r"-\d+$", "", stem) + suffix,
        re.sub(r"-\d+x\d+$", "", stem) + suffix,
    ]


def copy_product_images(urls: list[str]) -> dict[str, str]:
    PRODUCT_IMAGE_DIR.mkdir(parents=True, exist_ok=True)
    by_name = upload_index()
    used_names: Counter[str] = Counter()
    mapped: dict[str, str] = {}

    for url in sorted(set(urls)):
        source = None
        for candidate in url_candidates(url):
            source = by_name.get(candidate.lower())
            if source:
                break
        if not source:
            continue

        base = safe_asset_name(source.name)
        stem = Path(base).stem
        suffix = Path(base).suffix or source.suffix or ".jpg"
        used_names[base] += 1
        filename = base if used_names[base] == 1 else f"{stem}-{used_names[base]}{suffix}"
        target = PRODUCT_IMAGE_DIR / filename
        if not target.exists() or target.stat().st_size != source.stat().st_size:
            shutil.copy2(source, target)
        mapped[url] = f"/product-images/{filename}"

    return mapped


def display_title(row: dict[str, Any]) -> str:
    name = clean(row.get("Vārds"))
    brand = product_brand(row)
    pieces = name_parts(name)
    if len(pieces) >= 3 and pieces[0].lower() == brand.lower():
        return pieces[-1]
    if len(pieces) >= 2 and (pieces[0] in KNOWN_BRANDS or category_name(pieces[0]) != "Equipment"):
        return pieces[-1]
    return name


def name_parts(name: str) -> list[str]:
    return [part.strip() for part in re.split(r"\s*[–-]\s*", clean(name)) if part.strip()]


def product_brand(row: dict[str, Any]) -> str:
    explicit = clean(row.get("Zīmoli"))
    if explicit and explicit != "1":
        return explicit
    pieces = name_parts(clean(row.get("Vārds")))
    if pieces and pieces[0] in KNOWN_BRANDS:
        return pieces[0]
    return "Karatekas"


def product_category(row: dict[str, Any]) -> str:
    explicit = clean(row.get("Kategorija"))
    if explicit:
        return category_name(explicit)
    pieces = name_parts(clean(row.get("Vārds")))
    if len(pieces) >= 2:
        inferred = category_name(pieces[1])
        if inferred != "Equipment":
            return inferred
    if pieces:
        inferred = category_name(pieces[0])
        if inferred != "Equipment":
            return inferred
    return "Equipment"


def row_stock(row: dict[str, Any]) -> int:
    stock = int(round(number(row.get("Noliktavā"))))
    if stock <= 0 and clean(row.get("Ir noliktavā?")) == "1":
        return 1
    return max(0, stock)


def woo_rows() -> list[dict[str, Any]]:
    if not WOO_SOURCE.exists():
        return []
    with WOO_SOURCE.open(encoding="utf-8-sig", newline="") as handle:
        sample = handle.read(4096)
        handle.seek(0)
        dialect = csv.Sniffer().sniff(sample, delimiters=",;\t")
        return [
            row
            for row in csv.DictReader(handle, dialect=dialect)
            if row.get("Tips") in {"simple", "variable", "variation"}
            and clean(row.get("Publicēts")) != "0"
        ]


def import_woo_items() -> list[dict[str, Any]]:
    rows = woo_rows()
    if not rows:
        return []

    imported_at = datetime.now(timezone.utc).isoformat()
    parents_by_sku = {clean(row.get("SKU")): row for row in rows if row.get("Tips") == "variable"}
    children_by_parent: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for row in rows:
        if row.get("Tips") == "variation" and clean(row.get("Vecāks")):
            children_by_parent[clean(row.get("Vecāks"))].append(row)

    all_urls: list[str] = []
    for row in rows:
        all_urls.extend(image_urls(row))
    image_map = copy_product_images(all_urls)
    items: list[dict[str, Any]] = []
    seen_skus: Counter[str] = Counter()

    def add_item(parent: dict[str, Any], child: dict[str, Any], size: str, colour: str) -> None:
        brand = product_brand(parent)
        category = product_category(parent)
        title = display_title(parent)
        retail = number(child.get("Akcijas cena")) or number(child.get("Parastā cena:"))
        if retail <= 0:
            return
        club = round(retail * 0.78, 2)
        sku = unique_value(clean(child.get("SKU")) or clean(parent.get("SKU")) or title, seen_skus)
        parent_images = [image_map[url] for url in image_urls(parent) if url in image_map]
        child_images = [image_map[url] for url in image_urls(child) if url in image_map]
        images = (child_images or parent_images)[:8]
        description = strip_html(clean(parent.get("Īss apraksts")) or clean(parent.get("Apraksts")))

        items.append(
            {
                "source_sheet": "woocommerce-csv",
                "brand": brand,
                "product_group": clean(parent.get("Kategorija")) or category,
                "category": category,
                "sku": sku,
                "title": title,
                "description": description,
                "stock": row_stock(child),
                "size": clean(size),
                "colour": colour_name(colour),
                "club_price": club,
                "retail_price": round(retail, 2),
                "images": images,
                "image_urls": image_urls(parent),
                "imported_at": imported_at,
            }
        )

    for parent_sku, parent in parents_by_sku.items():
        children = children_by_parent.get(parent_sku, [])
        for child in children:
            attrs = parse_attrs(child)
            colours = attrs.get("color", [""])
            sizes = attrs.get("size", [""])
            add_item(parent, child, sizes[0] if sizes else "", colours[0] if colours else "")

    for row in rows:
        if row.get("Tips") != "simple":
            continue
        attrs = parse_attrs(row)
        colours = attrs.get("color", [""])
        sizes = attrs.get("size", [""])
        for colour in colours:
            for size in sizes:
                child = dict(row)
                suffix = "-".join(part for part in [colour, size] if part)
                if suffix:
                    child["SKU"] = f"{clean(row.get('SKU'))}-{slug(suffix)}"
                add_item(row, child, size, colour)

    return items


def build_products(items: list[dict[str, Any]]) -> tuple[list[str], list[dict[str, Any]]]:
    visible_items = [
        item
        for item in items
        if item["source_sheet"] == "woocommerce-csv"
        and (number(item["club_price"]) > 0 or number(item["retail_price"]) > 0)
    ]
    grouped: dict[tuple[str, str, str], list[dict[str, Any]]] = defaultdict(list)
    for item in visible_items:
        grouped[(item["brand"], item["category"], item["title"])].append(item)

    positions = [("0%", "0%"), ("50%", "0%"), ("100%", "0%"), ("0%", "100%"), ("50%", "100%"), ("100%", "100%")]
    categories = sorted(
        {item["category"] for item in visible_items},
        key=lambda category: (CATEGORY_ORDER.get(category, 999), category),
    )
    products: list[dict[str, Any]] = []
    seen_product_ids: Counter[str] = Counter()

    for index, ((brand, category, title), variations) in enumerate(grouped.items()):
        deduped: dict[tuple[str, str, float, float], dict[str, Any]] = {}
        for item in variations:
            key = (
                item["size"].lower(),
                item["colour"].lower(),
                round(number(item["club_price"]), 2),
                round(number(item["retail_price"]), 2),
            )
            current = deduped.get(key)
            if current is None or item["stock"] > current["stock"]:
                deduped[key] = item

        variations = sorted(
            deduped.values(),
            key=lambda item: (item["stock"] <= 0, item["colour"], item["size"], item["sku"]),
        )
        first = variations[0]
        product_id = unique_value(slug(f"{brand}-{title}"), seen_product_ids)
        product = {
            "id": product_id,
            "name": title,
            "brand": brand,
            "category": category,
            "description": first["description"],
            "specs": sorted({value for item in variations for value in [category, item["colour"]] if value})[:4],
            "tags": [brand, category],
            "sheetX": positions[index % len(positions)][0],
            "sheetY": positions[index % len(positions)][1],
            "images": first["images"],
            "variations": [],
        }

        for item in variations:
            club_price = item["club_price"] or item["retail_price"] * 0.78
            retail_price = item["retail_price"] or club_price * 1.28
            purchase = round(club_price * 0.62, 2)
            label_parts = [part for part in [item["colour"], item["size"]] if part]
            product["variations"].append(
                {
                    "id": slug(item["sku"]),
                    "sku": item["sku"],
                    "name": " / ".join(label_parts) or item["sku"],
                    "color": item["colour"],
                    "size": item["size"],
                    "b2c": round(retail_price, 2),
                    "b2b": round(club_price, 2),
                    "stock": {
                        "physical": item["stock"],
                        "reserved": 0,
                        "expected": 0,
                        "purchase": purchase,
                        "shipping": round(max(0.5, purchase * 0.05), 2),
                        "customs": round(max(0.15, purchase * 0.02), 2),
                        "vatRate": 21,
                        "fx": 1,
                        "lots": [
                            {
                                "batch": "WOO-KARATEKAS",
                                "qty": item["stock"],
                                "purchase": purchase,
                                "shipping": round(max(0.5, purchase * 0.05), 2),
                                "customs": round(max(0.15, purchase * 0.02), 2),
                                "vatRate": 21,
                                "fx": 1,
                                "eta": "на складе" if item["stock"] > 0 else "под заказ",
                            }
                        ],
                    },
                }
            )

        products.append(product)

    def product_available(product: dict[str, Any]) -> int:
        return sum(
            max(0, variation["stock"]["physical"] - variation["stock"]["reserved"])
            for variation in product["variations"]
        )

    products.sort(
        key=lambda p: (
            product_available(p) <= 0,
            CATEGORY_ORDER.get(p["category"], 999),
            p["brand"],
            p["name"],
        )
    )
    return categories, products


def write_ts(categories: list[str], products: list[dict[str, Any]]) -> None:
    content = (
        "// Generated from WooCommerce CSV and Excel price list by scripts/import_karate_catalog.py.\n"
        "// Do not edit product rows manually; update source files and rerun the importer instead.\n\n"
        "import type { Product } from './store-data';\n\n"
        f"export const karateCategories = {ts(categories)} as const;\n\n"
        f"export const karateProducts: Product[] = {ts(products)};\n"
    )
    TS_PATH.write_text(content, encoding="utf-8")


def import_rows() -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    excel_items = import_excel_rows()
    woo_items = import_woo_items()
    return excel_items + woo_items, woo_items


if __name__ == "__main__":
    all_items, woo_items = import_rows()
    categories, products = build_products(woo_items)
    write_ts(categories, products)
    print(
        json.dumps(
            {
                "imported_rows": len(all_items),
                "woocommerce_rows": len(woo_items),
                "products": len(products),
                "categories": categories,
                "brands": sorted({item["brand"] for item in woo_items}),
                "images_copied_to": str(PRODUCT_IMAGE_DIR),
                "ts": str(TS_PATH),
            },
            ensure_ascii=False,
            indent=2,
        )
    )
